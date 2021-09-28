import math
import os

from os.path import dirname

import numpy as np
import openpyxl
from numpy import sort, log, arange, exp
from pandas import read_excel

import data_poro

BASE_DIR = os.environ['BASE_DIR']

def convert_float(value):
    try:
        return float(value)
    except ValueError as verr:
        return False
    except Exception as ex:
        return False


class SheetReader:
    url = None
    Columns = []
    old_workbook = None
    sheet = None
    FZI = openpyxl.Workbook()
    FZI_sheet = None
    fzi = None

    def __init__(self, method, url, depth_column_name, porv_column_name, pron_column_name, layer_column_name,
                 note_column_name):
        self.depth_column_name = depth_column_name
        self.pron_column_name = pron_column_name
        self.porv_column_name = porv_column_name
        self.layer_column_name = layer_column_name
        self.note_column_name = note_column_name
        self.url = url

        # доступные методы: fzi, winland, lucia ...
        self.method = method
        self.old_workbook = openpyxl.load_workbook(url)
        self.sheet = self.old_workbook[self.old_workbook.sheetnames[0]]
        self.run()

    def run(self):
        self.pull_columns_to_new_workbook([self.depth_column_name, self.porv_column_name, self.pron_column_name])
        calc_method = getattr(self, self.method)
        calc_method()

    def pull_columns_to_new_workbook(self, columns):
        row_number_old_sheet = 2
        row_number_new_sheet = 2
        self.FZI_sheet = self.FZI.active

        # очень нужно не удаляй
        columns_name = ['Глубина', 'Пористость', 'Проницаемость', 'FZI']
        for i in range(len(columns_name)):
            cell = self.FZI_sheet.cell(row=1, column=i + 1)
            cell.value = columns_name[i]

        while (self.sheet[columns[0] + str(row_number_old_sheet)]).value:

            if self._row_valid(row_number_old_sheet):
                for i in range(len(columns)):
                    cell = self.FZI_sheet.cell(row=row_number_new_sheet, column=i + 1)
                    cell.value = self.sheet[columns[i] + str(row_number_old_sheet)].value
                row_number_new_sheet += 1
            row_number_old_sheet += 1
        self.valid_data = BASE_DIR + '/Files/valid_data.xlsx'
        self.FZI.save('Files/valid_data.xlsx')

    def _row_valid(self, row_number: int) -> bool:
        if self.sheet[self.depth_column_name + str(row_number)].value == \
                self.sheet[self.depth_column_name + str(row_number - 1)].value:
            return False

        if str(self.sheet[self.note_column_name + str(row_number)].value).__contains__('трещина'):
            return False

        if str(self.sheet[self.porv_column_name + str(row_number)].value) == '0':
            return False
        if str(self.sheet[self.porv_column_name + str(row_number)].value) == 'None':
            return False

        if str(self.sheet[self.pron_column_name + str(row_number)].value) == '0':
            return False
        if str(self.sheet[self.pron_column_name + str(row_number)].value) == 'None':
            return False

        for exception in ['ртинс', 'c1ok', 'c1bb', 'c1tl']:
            if self.sheet[self.layer_column_name + str(row_number)].value.__contains__(exception):
                return False

        return True

    def fzi(self):
        df = read_excel('Files/valid_data.xlsx')
        df['Пористость'] = df['Пористость'] / 100
        df['porosity*'] = df['Пористость'] / (1 - df['Пористость'])
        df['корень'] = (df['Проницаемость'] / df['Пористость']) ** 0.5
        df['RQI'] = 0.0314 * df['корень']
        df['FZI'] = df['RQI'] / df['porosity*']
        df['Log(FZI)'] = log(df['FZI'])
        df['probability'] = 1 / (len(df)) * (df.index + 1)
        df = df.sort_values(by='Log(FZI)', ascending=True)  # сортировка данных по возрастанию
        df['probability'] = sorted(df['probability'])

        self.file_method_data = 'Files/fzi.xlsx'
        df.to_excel(self.file_method_data)

    def lucia(self):
        df = read_excel('Files/valid_data.xlsx')
        df['Пористость'] = df['Пористость'] / 100
        df['Проницаемость'] = df['Проницаемость']

        self.file_method_data = 'Files/lucia.xlsx'
        df.to_excel(self.file_method_data)

    def winland(self):
        df = read_excel('Files/valid_data.xlsx')
        df['Пористость'] = df['Пористость'] / 100
        df['winland'] = 10 ** (.732 + .588 * log(df['Проницаемость']) - .865 * log(df['Пористость']))
        df['winland'] = log(df['winland'])
        df['probability'] = 1 / (len(df)) * (df.index + 1)
        df = df.sort_values(by='winland', ascending=True)  # сортировка данных по возрастанию
        df['probability'] = sorted(df['probability'])

        x = sort(df['winland'])
        y = arange(1, len(x) + 1) / len(x)

        self.file_method_data = 'Files/winland.xlsx'
        df.to_excel(self.file_method_data)

    @staticmethod
    def get_column(filename, column_name) -> []:
        df = openpyxl.load_workbook(filename).active
        column = []
        for i in range(len(df[column_name])):
            column.append(df[column_name + str(i + 1)].value)
        return column

    def get_column_method(self, column_name) -> []:
        df = read_excel(self.file_method_data)

        return df[column_name]


class Winland:
    def __init__(self, main_file, depth, porv, pron, layer, note, c_sw, r_sw):
        self.depth_column_name = depth
        self.porv_column_name = porv
        self.pron_column_name = pron
        self.layer_column_name = layer
        self.note_column_name = note

        self.current_sw = SheetReader.get_column(main_file, c_sw)
        self.residual_sw = SheetReader.get_column(main_file, r_sw)

        self.calc_main(main_file)

    def calc_main(self, filename):
        wb_fes = SheetReader('winland', filename, self.depth_column_name,
                             self.porv_column_name,
                             self.pron_column_name,
                             self.layer_column_name,
                             self.note_column_name)

        self.probability = wb_fes.get_column_method(['probability'])
        self.method = wb_fes.get_column_method(['winland'])
        self.pron_por = wb_fes.get_column_method(['Пористость', 'Проницаемость', 'winland'])

    def calc_rocktype(self, rock_type_borders):
        x_pts_modify = [-math.inf] + sorted(rock_type_borders)
        dots_rock_type = []
        self.modify_current_sw = []
        self.modify_residual_sw = []
        for x in range(len(x_pts_modify) - 1):
            rock_type_n = []
            modify_current_sw_n = []
            modify_residual_sw_n = []

            for i in range(len(self.pron_por['winland'])):
                if x_pts_modify[x] < self.pron_por['winland'][i] < x_pts_modify[x + 1]:
                    rock_type_n.append([self.pron_por['Пористость'][i], self.pron_por['Проницаемость'][i]])
                    value_current_sw = convert_float(self.current_sw[i])
                    value_residual_sw = convert_float(self.residual_sw[i])
                    if value_current_sw:
                        modify_current_sw_n.append(value_current_sw)
                    if value_residual_sw:
                        modify_residual_sw_n.append(value_residual_sw)

            dots_rock_type.append(rock_type_n)
            self.modify_current_sw.append(modify_current_sw_n)
            self.modify_residual_sw.append(modify_residual_sw_n)
        return dots_rock_type


class Lucia:
    def __init__(self, main_file, depth, porv, pron, layer, note, c_sw, r_sw):
        self.depth_column_name = depth
        self.porv_column_name = porv
        self.pron_column_name = pron
        self.layer_column_name = layer
        self.note_column_name = note

        self.current_sw_column_name = c_sw
        self.residual_sw_column_name = r_sw

        self.calc_main(main_file)
        self.calculate_RTWS(main_file)

        self.borders = [[data_poro.poro_05, data_poro.perm_05], [data_poro.poro_15, data_poro.perm_15],
                        [data_poro.poro_25, data_poro.perm_25], [data_poro.poro_4, data_poro.perm_4]]

    def calc_main(self, filename):
        wb_fes = SheetReader('lucia', filename, self.depth_column_name,
                             self.porv_column_name,
                             self.pron_column_name,
                             self.layer_column_name,
                             self.note_column_name)

        self.poro = wb_fes.get_column_method(['Пористость']).values.flatten()
        self.pron = wb_fes.get_column_method(['Проницаемость']).values.flatten()

    def calculate_RTWS(self, sw_file):
        current_sw = SheetReader.get_column(sw_file, 'D')
        residual_sw = SheetReader.get_column(sw_file, 'V')

        self.modify_current_sw = []
        for i in current_sw:
            if isinstance(i, float):
                self.modify_current_sw.append(i / 100)
        self.modify_current_sw = sorted(self.modify_current_sw)

        self.modify_residual_sw = []
        for i in residual_sw:
            if isinstance(i, float):
                self.modify_residual_sw.append(i / 100)
        self.modify_residual_sw = sorted(self.modify_residual_sw)

    def calc_rocktype(self):
        dots_rock_type = {'orange': [[], []], 'green': [[], []], 'blue': [[], []], 'red': [[], []], 'grey': [[], []]}

        for i in range(0, len(self.poro)):
            y1 = np.exp((22.56-12.08*np.log(4.0)) + ((8.671-3.603*np.log(4.0))*np.log(self.poro[i])))
            if self.pron[i] < y1:
                dots_rock_type['orange'][0].append(self.poro[i])
                dots_rock_type['orange'][1].append(self.pron[i])
                continue

            y2 = np.exp((22.56-12.08*np.log(2.5)) + ((8.671-3.603*np.log(2.5))*np.log(self.poro[i])))
            if self.pron[i] < y2:
                dots_rock_type['green'][0].append(self.poro[i])
                dots_rock_type['green'][1].append(self.pron[i])
                continue

            y3 = np.exp((22.56-12.08*np.log(1.5)) + ((8.671-3.603*np.log(1.5))*np.log(self.poro[i])))
            if self.pron[i] < y3:
                dots_rock_type['blue'][0].append(self.poro[i])
                dots_rock_type['blue'][1].append(self.pron[i])
                continue

            y4 = np.exp((22.56-12.08*np.log(0.5)) + ((8.671-3.603*np.log(0.5))*np.log(self.poro[i])))
            if self.pron[i] < y4:
                dots_rock_type['red'][0].append(self.poro[i])
                dots_rock_type['red'][1].append(self.pron[i])
                continue

            dots_rock_type['grey'][0].append(self.poro[i])
            dots_rock_type['grey'][1].append(self.pron[i])
        self.dots_rock_type = dots_rock_type
        return dots_rock_type


class Fzi:
    probability = []
    method = []

    rock_type_borders = []
    dots_rock_type = []
    pron_por = []

    rock_type_chart_scale = 'log'
    RTWS_chart_type = 'current'

    rock_type_colors = []

    def __init__(self, main_file, depth, porv, pron, layer, note, c_sw, r_sw):
        super().__init__()

        self.depth_column_name = depth
        self.porv_column_name = porv
        self.pron_column_name = pron
        self.layer_column_name = layer
        self.note_column_name = note

        self.current_sw = SheetReader.get_column(main_file, c_sw)
        self.residual_sw = SheetReader.get_column(main_file, r_sw)

        self.calc_main(main_file)

    def calc_main(self, filename):
        wb_fes = SheetReader('fzi', filename, self.depth_column_name,
                             self.porv_column_name,
                             self.pron_column_name,
                             self.layer_column_name,
                             self.note_column_name, )

        self.probability = wb_fes.get_column_method(['probability'])
        self.method = wb_fes.get_column_method(['Log(FZI)'])
        self.pron_por = wb_fes.get_column_method(['Пористость', 'Проницаемость', 'Log(FZI)'])

    def calc_rocktype(self, rock_type_borders):
        x_pts_modify = [-math.inf] + sorted(rock_type_borders)
        dots_rock_type = []
        self.modify_current_sw = []
        self.modify_residual_sw = []
        for x in range(len(x_pts_modify) - 1):
            rock_type_n = []
            modify_current_sw_n = []
            modify_residual_sw_n = []

            for i in range(len(self.pron_por['Log(FZI)'])):
                if x_pts_modify[x] < self.pron_por['Log(FZI)'][i] < x_pts_modify[x + 1]:
                    rock_type_n.append([self.pron_por['Пористость'][i], self.pron_por['Проницаемость'][i]])
                    value_current_sw = convert_float(self.current_sw[i])
                    value_residual_sw = convert_float(self.residual_sw[i])
                    if value_current_sw:
                        modify_current_sw_n.append(value_current_sw)
                    if value_residual_sw:
                        modify_residual_sw_n.append(value_residual_sw)

            dots_rock_type.append(rock_type_n)
            self.modify_current_sw.append(modify_current_sw_n)
            self.modify_residual_sw.append(modify_residual_sw_n)
        return dots_rock_type