import openpyxl
import pandas as pd
import numpy as np
from matplotlib import pyplot as plt


class SheetReader:
    url = None
    Columns = []
    old_workbook = None
    sheet = None
    FZI = openpyxl.Workbook()
    FZI_sheet = None
    fzi = None

    def __init__(self, method, url, depth_column_name, porv_column_name, pron_column_name, layer_column_name, note_column_name):
        self.depth_column_name = depth_column_name
        self.pron_column_name = pron_column_name
        self.porv_column_name = porv_column_name
        self.layer_column_name = layer_column_name
        self.note_column_name = note_column_name
        self.url = url

        # доступные методы: fzi, winland, ...
        self.method = method
        self.old_workbook = openpyxl.load_workbook(url)
        print(self.old_workbook)
        self.sheet = self.old_workbook[self.old_workbook.sheetnames[0]]
        self.run()

    def run(self):
        self.pull_columns_to_new_workbook([self.depth_column_name, self.porv_column_name, self.pron_column_name])
        calc_method = getattr(self, self.method)
        calc_method()

    def pull_columns_to_new_workbook(self, columns):
        row_number_old_sheet = 2
        row_number_new_sheet = 2
        self.FZI_sheet = self.FZI.active

        # очень нужно не удаляй
        columns_name = ['Глубина', 'Пористость', 'Проницаемость']
        for i in range(len(columns_name)):
            cell = self.FZI_sheet.cell(row=1, column=i + 1)
            cell.value = columns_name[i]

        while (self.sheet[columns[0] + str(row_number_old_sheet)]).value:

            if self._row_valid(row_number_old_sheet):
                for i in range(len(columns)):
                    cell = self.FZI_sheet.cell(row=row_number_new_sheet, column=i + 1)
                    cell.value = self.sheet[columns[i] + str(row_number_old_sheet)].value
                row_number_new_sheet += 1
            row_number_old_sheet += 1

        self.FZI.save('Files/valid_data.xlsx')

    def _row_valid(self, row_number: int) -> bool:
        if self.sheet[self.depth_column_name + str(row_number)].value == \
                self.sheet[self.depth_column_name + str(row_number - 1)].value:
            return False

        if str(self.sheet[self.note_column_name + str(row_number)].value).__contains__('трещина'):
            return False

        if str(self.sheet[self.porv_column_name + str(row_number)].value) == '0':
            return False
        if str(self.sheet[self.porv_column_name + str(row_number)].value) == 'None':
            return False

        if str(self.sheet[self.pron_column_name + str(row_number)].value) == '0':
            return False
        if str(self.sheet[self.pron_column_name + str(row_number)].value) == 'None':
            return False

        for exception in ['ртинс', 'c1ok', 'c1bb', 'c1tl']:
            if self.sheet[self.layer_column_name + str(row_number)].value.__contains__(exception):
                return False

        return True

    def fzi(self):
        df = pd.read_excel('Files/valid_data.xlsx')
        df['Пористость'] = df['Пористость'] / 100
        df['porosity*'] = df['Пористость'] / (1 - df['Пористость'])
        df['корень'] = (df['Проницаемость'] / df['Пористость']) ** 0.5
        df['RQI'] = 0.0314 * df['корень']
        df['FZI'] = df['RQI'] / df['porosity*']
        df['Log(FZI)'] = np.log(df['FZI'])
        df['probability'] = 1 / (len(df)) * (df.index+1)
        df = df.sort_values(by='Log(FZI)', ascending=True)  # сортировка данных по возрастанию
        df['probability'] = sorted(df['probability'])



        self.file_method_data = 'Files/fzi.xlsx'
        df.to_excel(self.file_method_data)

    def winland(self):
        df = pd.read_excel('Files/valid_data.xlsx')
        df['Пористость'] = df['Пористость'] / 100
        df['winland'] = 10 ** (.732+.588*np.log(df['Проницаемость'])-.865*np.log(df['Пористость']))
        df['winland'] = np.log(df['winland'])
        df['probability'] = 1 / (len(df)) * (df.index+1)
        df = df.sort_values(by='winland', ascending=True)  # сортировка данных по возрастанию
        df['probability'] = sorted(df['probability'])

        x = np.sort(df['winland'])
        y = np.arange(1, len(x) + 1) / len(x)
        plt.plot(x, y, marker='.', linestyle='none')
        plt.hist(x, bins=500)
        plt.xscale('log')

        plt.show()

        self.file_method_data = 'Files/winland.xlsx'
        df.to_excel(self.file_method_data)

    def lucia(self):
        print('lucia')

    @staticmethod
    def get_column(filename, column_name) -> []:
        df = openpyxl.load_workbook(filename).active
        column = []
        for i in range(len(df[column_name])):
            column.append(df[column_name + str(i+1)].value)
        return column

    def get_column_method(self, column_name) -> []:
        df = pd.read_excel(self.file_method_data)
        return df[column_name]
